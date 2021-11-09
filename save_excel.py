from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import numpy as np
import time

def Csv_save(inference_result, Date, csv_path, end_of_day):

    try:
        write_wb = load_workbook(csv_path + "/" + Date + ".xlsx")
        new = False
    except FileNotFoundError:
        write_wb = Workbook()
        write_wb.save(csv_path + "/" + Date + ".xlsx")
        write_wb = load_workbook(csv_path + "/" + Date + ".xlsx")
        new = True

    write_ws = write_wb.active

    if new == False:
        idx = write_ws.max_row + 1
    elif new == True:
        idx = 4
    
    row = np.arange(3, idx+1)
    column = [2, 3, 4, 5, 6, 7, 8]

    write_ws['B3'] = 'standard'
    write_ws['C3'] = 'diamond'
    write_ws['D3'] = 'jinro'
    write_ws['E3'] = 'chungha'
    write_ws['F3'] = 'beer'
    write_ws['G3'] = 'etc'
    write_ws['H3'] = 'empty'

    beer = 0
    etc = 0
    for i in inference_result:
        if i == "Standard":
            write_ws['B'+str(idx)] = inference_result[i]
        elif i == 'Diamond':
            write_ws['C'+str(idx)] = inference_result[i]
        elif i == "Jinro":
            write_ws['D'+str(idx)] = inference_result[i]
        elif i == "Chungha":
            write_ws['E'+str(idx)] = inference_result[i]
        elif i == "Cass":
            beer += inference_result[i]
        elif i == "Bud":
            beer += inference_result[i]
        elif i == "Cafri":
            beer += inference_result[i]
        elif i == "Maesil":
            etc += inference_result[i]
        elif i == 'Etc':
            etc += inference_result[i]
        elif i == 'Empty':
            write_ws['H'+str(idx)] = inference_result[i]
    write_ws['F' + str(idx)] = beer
    write_ws['G' + str(idx)] = etc


    write_ws['A' + str(idx)] = time.strftime('%H-%M-%S', time.localtime())


    for num_1 in row:
        for num_2 in column:
            write_ws.cell(row = num_1, column = num_2).alignment = Alignment(horizontal='center', vertical='center')

    if end_of_day==True:
        print('Final')
        sheet = write_wb.get_sheet_by_name('Sheet')
        if 'Final' in write_wb.sheetnames:
            write_ws2 = write_wb.get_sheet_by_name('Final')
        else:
            write_ws2 = write_wb.create_sheet('Final')

        write_ws2['B3'] = 'standard'
        write_ws2['C3'] = 'diamond'
        write_ws2['D3'] = 'jinro'
        write_ws2['E3'] = 'chungha'
        write_ws2['F3'] = 'beer'
        write_ws2['G3'] = 'etc'
        write_ws2['H3'] = 'empty'

        write_ws2['A4'] = time.strftime('%H-%M-%S', time.localtime())

        for i in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            count = 0
            for j in range(4, idx+1):
                count += sheet[i+str(j)].value

            write_ws2[i+str(4)] = count



    write_wb.save(csv_path + "/" + Date + ".xlsx")

def final_read(Date, csv_path, csv_path_final):
    write_wb = load_workbook(csv_path + "/" + Date + ".xlsx")

    write_ws = write_wb['Final']
    data = write_ws[4]

    _data =[]
    for i in range(len(list(data))):

        if i == 0 :
            _data.append(Date)
        else:
            _data.append(data[i].value)

    _data = np.array(_data).reshape(-1, 1)
    _data = _data.transpose()

    filepath = csv_path_final + "/" + Date +"_1.csv"
    np.savetxt(filepath, _data, delimiter = ',', fmt = '%s')

